import openpyxl


class HomePageData:

    #test_HomePage_data = [{"name":"eude", "email":"eude@eude.com", "gender":"Male"}, {"name":"yola", "email":"eudee@eude.com", "gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        elements = {}
        data = openpyxl.load_workbook("C:\\Users\\Eudyy\\Videos\\SeleniumFramework\\TestData\\Book3.xlsx")
        sheet = data.active
        for i in range(1, sheet.max_row +1):
            if sheet.cell(row= i, column=1).value == test_case_name:
                for j in range(1, sheet.max_column +1):
                    elements[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [elements]